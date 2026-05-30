"""
config/config_loader.py
Парсер файла config/config.xlsx для TegamCAL.

Читает три листа:
  - Technicians  → список техников (только Active=YES)
  - References   → 4 калибровочные точки с допусками и параметрами диапазона
  - Settings     → настройки программы (порты, задержки, пути)

Использование:
    from config.config_loader import load_config
    cfg = load_config()                   # читает ./config/config.xlsx
    cfg = load_config("path/to/config.xlsx")  # или явный путь

Возвращает объект CalibrationConfig с атрибутами:
    cfg.technicians   → список Technician
    cfg.points        → список CalPoint (4 калибровочные точки)
    cfg.settings      → dict настроек
"""

import os
from dataclasses import dataclass, field
from typing import List, Dict, Any

# openpyxl — единственная зависимость
try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("openpyxl not installed. Run: pip install openpyxl")


# ---------------------------------------------------------------
# Структура данных
# ---------------------------------------------------------------

@dataclass
class Technician:
    """Один сертифицированный техник из листа Technicians."""
    id: str
    name: str
    level: str
    signature: str


@dataclass
class CalPoint:
    """Одна калибровочная точка из листа References.

    Поля:
        point           — номер точки (1–4)
        relay           — команда реле ("R1"–"R4")
        nominal_ohm     — номинальное значение в Ом (для расчётов)
        ref_value_ohm   — реальное значение эталона в Ом (из свидетельства)
        tol_percent     — допуск в % от номинала
        tol_counts      — допуск в отсчётах (ед. разрешения)
        resolution      — цена деления шкалы (в Ом)
        range_code      — код диапазона Tegam 1750 (для команды R<n>X)
        display_unit    — единица отображения ("Ohm", "KOhm", "MOhm")
        notes           — примечание (например, "2-wire demo only")
    """
    point: int
    relay: str
    nominal_ohm: float
    ref_value_ohm: float
    tol_percent: float
    tol_counts: int
    resolution: float
    range_code: int
    display_unit: str
    notes: str

    def tolerance_ohm(self) -> float:
        """Абсолютный допуск в Ом по формуле Tegam.

        tolerance = nominal × (Tol_% / 100) + Tol_Counts × Resolution
        """
        return self.nominal_ohm * (self.tol_percent / 100.0) + self.tol_counts * self.resolution

    def is_pass(self, measured_ohm: float) -> bool:
        """PASS если отклонение от эталона не превышает допуск."""
        deviation = abs(measured_ohm - self.ref_value_ohm)
        return deviation <= self.tolerance_ohm()


@dataclass
class CalibrationConfig:
    """Полная конфигурация калибровки, загруженная из config.xlsx."""
    technicians: List[Technician] = field(default_factory=list)
    points: List[CalPoint] = field(default_factory=list)
    settings: Dict[str, Any] = field(default_factory=dict)

    def get_technician_names(self) -> List[str]:
        """Список имён активных техников для выпадающего списка в UI."""
        return [t.name for t in self.technicians]

    def get_point(self, point_num: int) -> CalPoint:
        """Получить CalPoint по номеру точки (1–4)."""
        for p in self.points:
            if p.point == point_num:
                return p
        raise ValueError(f"Calibration point {point_num} not found in config")

    def get_setting(self, key: str, default=None):
        """Получить настройку по ключу. Возвращает default если не найдено."""
        return self.settings.get(key, default)


# ---------------------------------------------------------------
# Вспомогательные функции
# ---------------------------------------------------------------

def _cell_value(ws, row: int, col: int):
    """Безопасное чтение ячейки — возвращает None если пусто."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        return val if val else None
    return val


def _to_float(val, label: str) -> float:
    """Конвертировать значение в float, с понятной ошибкой если не получается."""
    try:
        return float(val)
    except (TypeError, ValueError):
        raise ValueError(f"Cannot convert '{val}' to float for field '{label}'")


def _to_int(val, label: str) -> int:
    """Конвертировать значение в int."""
    try:
        return int(val)
    except (TypeError, ValueError):
        raise ValueError(f"Cannot convert '{val}' to int for field '{label}'")


# ---------------------------------------------------------------
# Парсеры листов
# ---------------------------------------------------------------

def _parse_technicians(ws) -> List[Technician]:
    """Парсить лист Technicians. Строка 1 — заголовки, строки 2+ — данные."""
    technicians = []
    # Ожидаемые колонки: ID, Name, Level, Signature, Active
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Пропускаем пустые строки
        if not row[0]:
            continue

        tech_id   = str(row[0]).strip()
        name      = str(row[1]).strip() if row[1] else ""
        level     = str(row[2]).strip() if row[2] else ""
        signature = str(row[3]).strip() if row[3] else ""
        active    = str(row[4]).strip().upper() if row[4] else "NO"

        # Включаем только активных техников (Active=YES)
        if active == "YES":
            technicians.append(Technician(
                id=tech_id,
                name=name,
                level=level,
                signature=signature,
            ))

    if not technicians:
        raise ValueError("No active technicians found in 'Technicians' sheet. "
                         "At least one technician must have Active=YES.")
    return technicians


def _parse_references(ws) -> List[CalPoint]:
    """Парсить лист References.
    Строка 1 — баннер-заголовок (пропускаем).
    Строка 2 — названия колонок (пропускаем).
    Строки 3–6 — 4 калибровочные точки.
    """
    points = []
    # Данные начинаются с row=3 (1=баннер, 2=заголовки)
    for row_idx in range(3, 8):
        point_num  = _cell_value(ws, row_idx, 1)
        if point_num is None:
            continue  # пустая строка после данных

        relay      = _cell_value(ws, row_idx, 2)
        nominal    = _cell_value(ws, row_idx, 3)
        ref_val    = _cell_value(ws, row_idx, 4)
        tol_pct    = _cell_value(ws, row_idx, 5)
        tol_counts = _cell_value(ws, row_idx, 6)
        resolution = _cell_value(ws, row_idx, 7)
        range_code = _cell_value(ws, row_idx, 8)
        disp_unit  = _cell_value(ws, row_idx, 9)
        notes      = _cell_value(ws, row_idx, 10) or ""

        points.append(CalPoint(
            point         = _to_int(point_num,  "Point"),
            relay         = str(relay),
            nominal_ohm   = _to_float(nominal,    "Nominal (Ohm)"),
            ref_value_ohm = _to_float(ref_val,    "Ref Value (Ohm)"),
            tol_percent   = _to_float(tol_pct,    "Tol_%"),
            tol_counts    = _to_int(tol_counts,   "Tol_Counts"),
            resolution    = _to_float(resolution, "Resolution"),
            range_code    = _to_int(range_code,   "Range Code"),
            display_unit  = str(disp_unit),
            notes         = str(notes),
        ))

    if len(points) != 4:
        raise ValueError(f"Expected 4 calibration points in 'References' sheet, got {len(points)}")
    return points


def _parse_settings(ws) -> Dict[str, Any]:
    """Парсить лист Settings.
    Строка 1 — баннер (пропускаем).
    Строка 2 — заголовки (пропускаем).
    Строки 3+ — пары Ключ / Значение.
    """
    settings = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        key = row[0]
        val = row[1]
        if not key:
            continue
        key = str(key).strip()
        # Пропускаем строки-легенды и комментарии
        if not key or key.startswith("YELLOW") or key.startswith("#"):
            continue
        settings[key] = val

    return settings


# ---------------------------------------------------------------
# Основная функция загрузки
# ---------------------------------------------------------------

def load_config(path: str = None) -> CalibrationConfig:
    """Загрузить конфигурацию из config.xlsx.

    Аргументы:
        path — путь к файлу. По умолчанию ./config/config.xlsx
               (относительно рабочей директории при запуске main.py).

    Возвращает:
        CalibrationConfig — заполненный объект конфигурации.

    Исключения:
        FileNotFoundError — файл не найден.
        ValueError        — ошибка формата данных в файле.
    """
    if path is None:
        # Путь по умолчанию — относительно корня проекта
        path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                            "config", "config.xlsx")

    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Config file not found: {path}\n"
            "Run make_config.py once to generate the template."
        )

    # data_only=True — читать вычисленные значения ячеек, не формулы
    wb = load_workbook(path, data_only=True)

    # Проверяем что все нужные листы на месте
    required_sheets = {"Technicians", "References", "Settings"}
    missing = required_sheets - set(wb.sheetnames)
    if missing:
        raise ValueError(f"Missing sheets in config.xlsx: {missing}")

    config = CalibrationConfig(
        technicians = _parse_technicians(wb["Technicians"]),
        points      = _parse_references(wb["References"]),
        settings    = _parse_settings(wb["Settings"]),
    )

    return config


# ---------------------------------------------------------------
# Быстрый тест при запуске напрямую
# ---------------------------------------------------------------
if __name__ == "__main__":
    import sys

    # Запуск: python config/config_loader.py [путь к config.xlsx]
    path = sys.argv[1] if len(sys.argv) > 1 else None

    print("=== TegamCAL Config Loader — self test ===\n")
    try:
        cfg = load_config(path)
    except FileNotFoundError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    print("--- Technicians ---")
    for t in cfg.technicians:
        print(f"  [{t.id}] {t.name} ({t.level})")

    print("\n--- Calibration Points ---")
    for p in cfg.points:
        tol = p.tolerance_ohm()
        print(f"  Point {p.point} | {p.relay} | "
              f"Nominal: {p.nominal_ohm} Ohm | "
              f"Ref: {p.ref_value_ohm} Ohm | "
              f"Tolerance: ±{tol:.6f} Ohm | "
              f"Range: R{p.range_code} | "
              f"Display: {p.display_unit}")

    print("\n--- Settings ---")
    for k, v in cfg.settings.items():
        print(f"  {k:30s} = {v}")

    print("\n=== OK ===")
